import ofxparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import io

def remove_closing_tags(content):
    content_no_closing_tags = re.sub(r'</CODE>', '', content)
    content_no_closing_tags = re.sub(r'</SEVERITY>', '', content_no_closing_tags)
    return content_no_closing_tags

def processar_arquivo_ofx(ofx, excel_filename='extrato.xlsx'):
    # 1. Informações da Conta
    account_info = {
        "Nome da Agência": [ofx.account.branch_id],
        "Número da Conta": [ofx.account.account_id],
        "Tipo de Conta": [ofx.account.account_type]
    }
    df_account_info = pd.DataFrame(account_info)

    # 2. Informações da Declaração
    statement_info = {
        "Data de Início": [ofx.account.statement.start_date.date()],
        "Data de Fim": [ofx.account.statement.end_date.date()]
    }
    df_statement_info = pd.DataFrame(statement_info)

    # 3. Separação de Entradas e Saídas
    entradas = []
    saidas = []

    for transaction in ofx.account.statement.transactions:
        trans_data = {
            'Data': transaction.date.date(),
            'Valor': float(transaction.amount),
            'Descrição': transaction.memo,
            'ID': transaction.id
        }
        if transaction.amount > 0:
            entradas.append(trans_data)
        else:
            saidas.append(trans_data)

    # Converte as listas de entradas e saídas para DataFrames do pandas
    df_entradas = pd.DataFrame(entradas)
    df_saidas = pd.DataFrame(saidas)

    # Salva todas as informações no Excel
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Escreve as informações da conta
        df_account_info.to_excel(writer, sheet_name='Extrato', startrow=2, startcol=0, index=False)
        df_statement_info.to_excel(writer, sheet_name='Extrato', startrow=2, startcol=4, index=False)

        # Adiciona títulos para entradas e saídas
        writer.sheets['Extrato'].cell(row=7, column=1).value = "Entradas"
        writer.sheets['Extrato'].cell(row=7, column=6).value = "Saídas"

        # Escreve as entradas e saídas
        df_entradas.to_excel(writer, sheet_name='Extrato', startrow=8, startcol=0, index=False)
        df_saidas.to_excel(writer, sheet_name='Extrato', startrow=8, startcol=5, index=False)

    # Carrega o arquivo Excel gerado para aplicar formatação adicional
    wb = load_workbook(excel_filename)
    ws = wb['Extrato']

    # 1. Adicionar título ao topo da planilha
    data_inicio = statement_info['Data de Início'][0].strftime('%Y-%m-%d')
    data_fim = statement_info['Data de Fim'][0].strftime('%Y-%m-%d')

    ws.insert_rows(0, amount=2)
    ws['A1'] = f"Extrato Bancário - {data_inicio} a {data_fim}"
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 2. Unificar e destacar as informações da conta
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A3:I4')  # Mescla as células para unificar as informações da conta
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # 3. Formatação das colunas
    ws.column_dimensions['A'].width = 15  # Data
    ws.column_dimensions['B'].width = 15  # Valor
    ws.column_dimensions['C'].width = 55  # Descrição
    ws.column_dimensions['D'].width = 25  # ID
    ws.column_dimensions['E'].width = 17  # Data de Início
    ws.column_dimensions['F'].width = 17  # Data de Fim
    ws.column_dimensions['G'].width = 17  # Valor
    ws.column_dimensions['H'].width = 55  # Descrição
    ws.column_dimensions['I'].width = 25  # ID

    # Formatação dos valores como moeda
    for col in ['B', 'G']:
        for row in range(9, ws.max_row + 1):
            cell = ws[f'{col}{row}']
            cell.number_format = '#,##0.00'

    # 4. Cores alternadas nas linhas de "Entradas" e "Saídas"
    fill_odd = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in range(9, ws.max_row + 1):
        if row % 2 == 0:  # Se a linha for par
            for col in range(1, 5):  # Colunas de A a D (Entradas)
                ws.cell(row=row, column=col).fill = fill_odd
            for col in range(6, 10):  # Colunas de F a H (Saídas)
                ws.cell(row=row, column=col).fill = fill_odd

    # 5. Adicionar bordas nas tabelas de "Entradas" e "Saídas"
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in range(9, ws.max_row + 1):
        for col in range(1, 5):  # Colunas de A a D (Entradas)
            ws.cell(row=row, column=col).border = thin_border
        for col in range(6, 10):  # Colunas de F a H (Saídas)
            ws.cell(row=row, column=col).border = thin_border

    # 7. Adicionar sumário dos totais no final de cada tabela (Entradas e Saídas)
    entradas_total_row = ws.max_row + 2
    saidas_total_row = ws.max_row + 2

    ws[f'A{entradas_total_row}'] = "Total Entradas:"
    ws[f'A{entradas_total_row}'].font = Font(bold=True)
    ws[f'B{entradas_total_row}'] = f"=SUM(B9:B{entradas_total_row-2})"
    ws[f'B{entradas_total_row}'].font = Font(bold=True)
    ws[f'B{entradas_total_row}'].number_format = '#,##0.00 €'

    ws[f'F{saidas_total_row}'] = "Total Saídas:"
    ws[f'F{saidas_total_row}'].font = Font(bold=True)
    ws[f'G{saidas_total_row}'] = f"=SUM(G9:G{saidas_total_row-2})"
    ws[f'G{saidas_total_row}'].font = Font(bold=True)
    ws[f'G{saidas_total_row}'].number_format = '#,##0.00'

    # Salva as mudanças
    wb.save(excel_filename)

    print("Arquivo Excel 'extrato.xlsx' criado e formatado com sucesso.")
